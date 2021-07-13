# Author - Leonardo Duarte Malta de Abreu 
# Github - leonardodma

# Imports Bibliotecas Básicas
from re import search
import re
from numpy.lib.function_base import append
import pandas as pd 
import numpy as np
import os, os.path
from pathlib import Path
import openpyxl
import easygui
from tqdm import tqdm

# Importe de Módulos
from crawler_viva_real import *
from crawler_zap_imoveis import *


class Crawler_Full():
    def __init__(self):
        # Variável para guardar os dados coletados
        self.all_data = None

        # Planilha selecionada 
        print('SELECIONE A PLANILHA COM O LAUDO QUE VOCÊ ESTÁ COLHENDO OS DADOS')
        self.laudo = str(self.get_file_path())
        print(f'VOCÊ SELECIONOU O LAUDO: {self.laudo}')

        # Chrome Driver
        chromedriver_path = Path(str(Path(__file__).parent.resolve()) + '\software\chromedriver_win32\chromedriver.exe')
        options = Options()
        options.add_argument("--window-size=1920x1800")
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.driver = webdriver.Chrome(executable_path = chromedriver_path, options=options)

    
    def get_file_path(self):
        #path = str(r'R:\Empírica Cobranças e Garantias\5 - Avaliações de Imóveis\Projeto estágio de férias\vba.txt')
        #print(f'O arquivo contendo o path do laudo é: {path}')

        #file = open(path, encoding="ascii").read()

        #return str(r'R:\Empírica Cobranças e Garantias\5 - Avaliações de Imóveis\Laudos Creditas\7.21\TESTE - ESTAGIO DE FÉRIAS - AUTOMATIZAÇÃO LAUDOS\07.21 - Maria José Pereira Dias.xlsm')

        return easygui.fileopenbox()


    def get_search_string(self):
        # Give the location of the file
        path = self.laudo
        wb_obj = openpyxl.load_workbook(path, data_only=True)
        self.ws = wb_obj['Modelo de Laudo']

        bairro = self.ws['E23'].value.strip()
        municipio = self.ws['E24'].value.strip()
        tipo = self.ws['E27'].value.strip()

        search_string = f'{bairro}, {municipio}'.strip()

        return search_string, tipo


    def merge_data(self, data_dict_list):
        frames = []

        for i in range(len(data_dict_list)):
            df = pd.DataFrame.from_dict(data_dict_list[i])
            frames.append(df)

        self.all_data = pd.concat(frames)


    def get_data(self):
        search_string, tipo_imovel = self.get_search_string()

        crawler_viva_real = Crawler_VivaReal(self.driver)
        crawler_zap_imoveis = Crawler_ZapImoveis(self.driver)

        viva_real = crawler_viva_real.get_data(search_string, tipo_imovel)
        zap_imoveis = crawler_zap_imoveis.get_data(search_string, tipo_imovel)

        self.merge_data([viva_real, zap_imoveis])


    def clean_dataframe(self):
        area_imovel = self.ws['U34'].value
        print(f'Área do imóvel analisado: {area_imovel}')

        print('LIMPANDO IMÓVEIS PELA ÁREA')
        self.all_data = self.all_data[(self.all_data['Área'] < area_imovel+20) & 
                                      (self.all_data['Área'] > area_imovel-20)]

        print('LIMPANDO ENDEREÇOS VAZIOS')
        self.all_data['Endereço'].replace('', np.nan, inplace=True)
        self.all_data.dropna(subset=['Endereço'], inplace=True)

        print('LIMPANDO ITENS REPETIDOS')
        self.all_data.drop_duplicates(subset=['Bairro', 'Endereço', 'Área', 'Preço'], keep='last')


    def save_dataframe(self):
        file_path = self.laudo
        barra = str(r" \ ")[1]
        save_path = barra.join(file_path.split(barra)[0:-1])+barra+'dados_coletados.xlsx'
        print('\n')
        print(f'Dados coletados foram salvos em: {save_path}')
        
        self.all_data.reset_index(drop=True, inplace=True)
        self.all_data.to_excel(save_path, sheet_name='Sheet1')


if __name__ == '__main__':
    crawler = Crawler_Full()
    crawler.get_data()
    crawler.clean_dataframe()
    crawler.save_dataframe()
