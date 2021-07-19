# Author - Leonardo Duarte Malta de Abreu 
# Github - leonardodma

from bs4 import BeautifulSoup
import requests
from fake_useragent import UserAgent
from lxml import etree
import pathlib
from pathlib import Path
import openpyxl

ua = UserAgent()
with open(str(pathlib.Path(__file__).parent.resolve())+str(r'\file.txt'), 'r') as f:
    dados = f.read().strip().split("\n")
    file = dados[0]
    amostra = dados[1]
    tentativas = int(dados[2])


def get_address():
    wb_obj = openpyxl.load_workbook(file, data_only=True)
    ws = wb_obj['Pesquisa de Mercado']

    endereco = ws['Q7'].value.strip()
    bairro = ws['Q8'].value.strip()

    search_string = f'{endereco} {bairro}'.strip()

    return search_string


def get_url(search_str):
    query = search_str.split()
    query='+'.join(query)
    url="https://www.google.co.in/search?q="+query+"&source=lnms&tbm=isch"

    print(url)

    return url


def get_dom(url):
    session = requests.Session()
    page = session.get(url, headers={"User-Agent": str(ua.chrome)})
    soup = BeautifulSoup(page.content, 'html.parser')
    dom = etree.HTML(str(soup))

    return dom


def download_image(src):
    path = str(pathlib.Path(__file__).parent.resolve()) +  str(Path(f'\img\\amostra{amostra}.png'))
    print('Imagem salva em:')
    print(path)
    f = open(path,'wb')
    f.write(requests.get(src).content)
    f.close()


def get_image(tentativas):
    url = get_url(get_address())
    dom = get_dom(url)
    xpath = ".//img[contains(@class, 't0fcAb')]"
    src = dom.xpath(xpath)[tentativas].items()[2][1]
    download_image(src)


if __name__ == '__main__':
    get_image(tentativas)
