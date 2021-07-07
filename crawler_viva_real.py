# Author - Leonardo Duarte Malta de Abreu 
# Github - leonardodma

# Imports Bibliotecas Básicas
import types
from numpy.lib.function_base import append
import pandas as pd 
import numpy as np
import time
import os, os.path
from pathlib import Path
import openpyxl

# Imports Selenium (Navegador Web) - Obter Links dos imóveis 
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select

# Imports Web Scraper - Obter Base de dados 
from bs4 import BeautifulSoup
from lxml import etree
import requests
import warnings
from fake_useragent import UserAgent
from requests.packages.urllib3.exceptions import InsecureRequestWarning


# Importe de Módulos
from Parseador import *


class Crawler_VivaReal():
    def __init__(self):
        # Iniciando Softwares Necessários
        # 1) Chrome Driver
        chromedriver_path = Path(str(Path(__file__).parent.resolve()) + '\software\chromedriver_win32\chromedriver.exe')
        options = Options()
        options.add_argument("--window-size=1920x1800")
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.driver = webdriver.Chrome(executable_path = chromedriver_path, options=options)

        # 2) Dom Xpath
        warnings.simplefilter('ignore',InsecureRequestWarning)
        self.ua = UserAgent()

        # Variável para guardar os dados coletados
        self.all_data = {}


    def get_search_string(self):
        # Give the location of the file
        path = "modelo_laudo.xlsm"
        wb_obj = openpyxl.load_workbook(path, data_only=True)
        self.ws = wb_obj['Modelo de Laudo']

        bairro = self.ws['E23'].value.strip()
        municipio = self.ws['E24'].value.strip()
        #tipo = self.ws['E27'].value.strip()
        tipo = 'Casa'

        search_string = f'{bairro} - {municipio}'

        return search_string, tipo


    def seleciona_tipo(self, tipo_imovel):
        if tipo_imovel == "Apartamento":
            self.driver.find_element_by_xpath('//*[@id="unit-type-APARTMENT|UnitSubType_NONE,DUPLEX,LOFT,STUDIO,TRIPLEX|RESIDENTIAL|APARTMENT"]').click()
        elif tipo_imovel == "Casa":
            self.driver.find_element_by_xpath('//*[@id="unit-type-HOME|UnitSubType_NONE,SINGLE_STOREY_HOUSE,VILLAGE_HOUSE,KITNET|RESIDENTIAL|HOME"]').click()
        elif tipo_imovel == "Sala":
            self.driver.find_element_by_xpath('//*[@id="unit-type-OFFICE|UnitSubType_NONE,OFFICE,FLOOR|COMMERCIAL|OFFICE"]').click()
        elif tipo_imovel == "Loja":
            self.driver.find_element_by_xpath('//*[@id="unit-type-HOME|UnitSubType_NONE|COMMERCIAL|COMMERCIAL_PROPERTY"]').click()
        elif tipo_imovel == "Loteamento":
            self.driver.find_element_by_xpath('//*[@id="unit-type-ALLOTMENT_LAND|UnitSubType_NONE,CONDOMINIUM,VILLAGE_HOUSE|RESIDENTIAL|RESIDENTIAL_ALLOTMENT_LAND"]').click()


    def properties_url(self):
        search_str, tipo_imovel = self.get_search_string()

        self.driver.get('https://www.vivareal.com.br/venda/')
        edit_filters = self.driver.find_element_by_xpath('//*[@id="js-site-main"]/div[2]/div[1]/section/header/div/div/div[3]/div/button')
        edit_filters.click()

        # Aceitar os cookies da página 
        time.sleep(2)
        try:
            accept_cookies = self.driver.find_element_by_xpath('//*[@id="cookie-notifier-cta"]')
            accept_cookies.click()
        except:
            pass

        # Colocar string de pesquisa 
        search_box = self.driver.find_element_by_xpath('//*[@id="filter-location-search-input"]')
        search_box.send_keys(search_str)
        
        # Esperar tempo para aparecer as sugestões
        time.sleep(3)

        # Selecionar a primeira sugestão 
        search_box.send_keys(Keys.RETURN)

        # Filtrar tipo de imóvel (Apartamento, Casa, Sala, Loja, Loteamento)
        # 1) Clicar em "mostrar todos"
        time.sleep(1)
        mostrar_todos = self.driver.find_element_by_xpath('//*[@id="js-site-main"]/div[2]/div[1]/nav/div/div/form/fieldset[1]/div[3]/div')
        mostrar_todos.click()

        # 2) Ajuste à janela de seleção do tipo de imóvel
        time.sleep(1)
        janela_selecao = self.driver.find_element_by_xpath('//*[@id="js-site-main"]/div[2]/div[1]/nav/div/div/form/fieldset[1]/div[3]/div/div/div')
        self.driver.execute_script("arguments[0].scrollIntoView();", janela_selecao)

        # 3) Selecionar Tipo de Imóvel
        time.sleep(1)
        self.seleciona_tipo(tipo_imovel)

        # 4) Reajustar à "mostrar todos" novamente
        time.sleep(1)
        self.driver.execute_script("arguments[0].scrollIntoView();", mostrar_todos)
        mostrar_todos.click()
        
        # Terminar a edição 
        time.sleep(1)
        finish_edit = self.driver.find_element_by_xpath('//*[@id="js-site-main"]/div[2]/div[1]/nav/div/div/form/div/a')
        finish_edit.click()

        # Retornar o URL com os imóveis 
        time.sleep(2)


        return self.driver.current_url
    

    def qtd_anuncios(self, dom):
        qtd = 1
        idx_li = 1
        final = False

        while not final:
            try:
                dom.xpath(f'/html/body/main/div[2]/div[1]/section/div[2]/div[1]/div[{idx_li}]')[0]
                qtd += 1
                idx_li += 1
            except:
                final = True
        
        return qtd


    def get_data(self):
        main_URL = self.properties_url()
        time.sleep(2)
        self.driver.quit()

        tipo = self.get_search_string()[1]

        estados = []
        cidades = []
        bairros = []
        enderecos = []
        areas = []
        quartos = []
        banheiros = [] 
        precos = []
        links = []

        end = False
        pagina = 0

        while not end:
            if pagina == 0:
                URL = main_URL
                print('')
            else:
                URL = main_URL+f'?pagina={pagina}'

            print(URL)
            session = requests.Session()
            print('Sessão Iniciada')
            page = session.get(URL, headers={"User-Agent": str(self.ua.chrome)})
            print('Conexão com o site feita com sucesso')
            print('\n')

            if page.status_code == 200:
                soup = BeautifulSoup(page.content, 'html.parser')
                dom = etree.HTML(str(soup))
                qtd_anuncios = self.qtd_anuncios(dom)

                for i in range(1, qtd_anuncios):
                    # Extrações 
                    estado, cidade, bairro, endereco = get_adress(dom, i)
                    area = get_area(dom, i)
                    quarto = get_quartos(dom, i)
                    banheiro = get_banheiros(dom, i)
                    preco = get_preco(dom, i)
                    link = 'https://www.vivareal.com.br' + get_link(dom, i)

                    # Adicionando extrações à lista
                    if len(preco) > 0:
                        estados.append(estado)
                        cidades.append(cidade)
                        bairros.append(bairro)
                        enderecos.append(endereco)
                        areas.append(area)
                        quartos.append(quarto)
                        banheiros.append(banheiro)
                        precos.append(preco)
                        links.append(link)

                next_page = get_next_page(dom, tipo)
                print(next_page)
                if next_page == '#pagina=':
                    end = True
                    
            pagina += 1

        self.all_data['Estado'] = estados
        self.all_data['Cidade'] = cidades
        self.all_data['Bairro'] = bairros
        self.all_data['Endereço'] = enderecos
        self.all_data['Área'] = areas
        self.all_data['Quartos'] = quartos
        self.all_data['Banheiros'] = banheiros
        self.all_data['Preço'] = precos
        self.all_data['Link'] = links
        

    def remove_by_idx(self, idxs):
        self.all_data['Estado'] = [i for n, i in enumerate(self.all_data['Estado']) if n not in idxs]
        self.all_data['Cidade'] = [i for n, i in enumerate(self.all_data['Cidade']) if n not in idxs]
        self.all_data['Bairro'] = [i for n, i in enumerate(self.all_data['Bairro']) if n not in idxs]
        self.all_data['Endereço'] = [i for n, i in enumerate(self.all_data['Endereço']) if n not in idxs]
        self.all_data['Área'] = [i for n, i in enumerate(self.all_data['Área']) if n not in idxs]
        self.all_data['Quartos'] = [i for n, i in enumerate(self.all_data['Quartos']) if n not in idxs]
        self.all_data['Banheiros'] = [i for n, i in enumerate(self.all_data['Banheiros']) if n not in idxs]
        self.all_data['Preço'] = [i for n, i in enumerate(self.all_data['Preço']) if n not in idxs]
        self.all_data['Link'] = [i for n, i in enumerate(self.all_data['Link']) if n not in idxs]


    def filter_by_area(self):
        area_imovel = self.ws['U34'].value
        print(f'Área do imóvel analisado: {area_imovel}')

        areas = self.all_data['Área']
        array_areas = []
        for area in areas:
            try:
                array_areas.append(int(area))
            except:
                x = int(area.split('-')[0].strip())
                y = int(area.split('-')[0].strip())
                array_areas.append((x+y)/2)

        idx_remove = []
        for i in range(len(array_areas)):
            if array_areas[i] < area_imovel-20 or array_areas[i] > area_imovel+20:
                idx_remove.append(i)
        
        return idx_remove


    def clean_dataframe(self):
        # Filtra Área
        self.remove_by_idx(self.filter_by_area())
    

    def save_dataframe(self):
        pd.DataFrame.from_dict(self.all_data).to_excel('imoveis_coletados.xlsx', sheet_name='Sheet1')


if __name__ == '__main__':
    crawler = Crawler_VivaReal()
    crawler.get_data()
    crawler.clean_dataframe()
    crawler.save_dataframe()
